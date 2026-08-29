import csv
import json
import re
from enum import StrEnum
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

import duckdb
from openpyxl import load_workbook
from pydantic import Field

from app.contracts.data import StrictModel


class ProfileLimits(StrictModel):
    max_bytes: int = Field(default=10_000_000, gt=0)
    max_rows: int = Field(default=10_000, gt=0)
    max_columns: int = Field(default=500, gt=0)
    max_header_length: int = Field(default=256, gt=0)
    xlsx_max_members: int = Field(default=2_000, gt=0)
    xlsx_max_total_uncompressed: int = Field(default=100_000_000, gt=0)
    xlsx_max_member_uncompressed: int = Field(default=25_000_000, gt=0)
    xlsx_max_compression_ratio: float = Field(default=100.0, gt=0)


class PrivacyRisk(StrEnum):
    LOW = "LOW"
    MEDIUM = "MEDIUM"
    HIGH = "HIGH"


class ColumnProfile(StrictModel):
    name: str
    type: str
    null_rate: float = Field(ge=0, le=1)
    distinct_estimate: int = Field(ge=0)


class PrivacyFinding(StrictModel):
    column: str
    category: str
    risk: PrivacyRisk
    detected_by: tuple[str, ...]


class SchemaProfile(StrictModel):
    format: str
    columns: tuple[ColumnProfile, ...]
    row_estimate: int = Field(ge=0)
    encoding: str | None = None
    delimiter: str | None = None
    sheets: tuple[str, ...] = ()
    candidate_keys: tuple[str, ...] = ()
    privacy_findings: tuple[PrivacyFinding, ...] = ()
    warnings: tuple[str, ...] = ()


_PII_TOKENS: dict[str, tuple[str, PrivacyRisk]] = {
    "cpf": ("BRAZILIAN_CPF", PrivacyRisk.HIGH),
    "email": ("EMAIL", PrivacyRisk.HIGH),
    "mail": ("EMAIL", PrivacyRisk.HIGH),
    "telefone": ("PHONE", PrivacyRisk.HIGH),
    "tel": ("PHONE", PrivacyRisk.HIGH),
    "celular": ("PHONE", PrivacyRisk.HIGH),
    "fone": ("PHONE", PrivacyRisk.HIGH),
    "nome": ("PERSON_NAME", PrivacyRisk.MEDIUM),
}
_EMAIL = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")
_CPF = re.compile(r"^\d{3}\.?\d{3}\.?\d{3}-?\d{2}$")
_PHONE = re.compile(r"^(?:\+?55\s?)?(?:\(?\d{2}\)?\s?)?9?\d{4}[-\s]?\d{4}$")
_PATTERNS = (("EMAIL", _EMAIL), ("BRAZILIAN_CPF", _CPF), ("PHONE", _PHONE))


def _normalized(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.casefold()).strip("_")


def _value_category(value: str) -> str | None:
    text = value.strip()
    return next((category for category, pattern in _PATTERNS if pattern.fullmatch(text)), None)


class SchemaProfiler:
    def __init__(self, root: Path, limits: ProfileLimits | None = None) -> None:
        self._root = root.resolve(strict=True)
        self._limits = limits or ProfileLimits()

    def _resolve(self, relative: str | Path) -> Path:
        requested = Path(relative)
        if requested.is_absolute() or ".." in requested.parts:
            raise ValueError("path escapes profiling root")
        unresolved = self._root / requested
        if any(
            part.is_symlink()
            for part in [unresolved, *unresolved.parents]
            if part != self._root.parent
        ):
            raise ValueError("symlinks are not accepted for profiling")
        try:
            candidate = unresolved.resolve(strict=True)
        except (FileNotFoundError, OSError) as exc:
            raise ValueError("profile target does not exist") from exc
        if candidate == self._root or self._root not in candidate.parents:
            raise ValueError("path escapes profiling root")
        if not candidate.is_file():
            raise ValueError("profile target must be a regular confined file")
        if candidate.stat().st_size > self._limits.max_bytes:
            raise ValueError("file exceeds profiling byte limit")
        return candidate

    def profile(self, relative: str | Path) -> SchemaProfile:
        path = self._resolve(relative)
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return self._profile_csv(path)
        if suffix in {".json", ".jsonl", ".ndjson"}:
            return self._profile_json(path, lines=suffix != ".json")
        if suffix == ".parquet":
            return self._profile_parquet(path)
        if suffix == ".xlsx":
            return self._profile_xlsx(path)
        raise ValueError("unsupported profile format")

    @staticmethod
    def _finding(
        name: str, values: list[Any], safe_name: str | None = None
    ) -> PrivacyFinding | None:
        normalized = _normalized(name)
        detected: list[str] = []
        category: str | None = None
        risk = PrivacyRisk.HIGH
        tokens = set(normalized.split("_"))
        header_category = _value_category(name)
        if header_category is not None:
            category = header_category
            risk = PrivacyRisk.HIGH
            detected.append("COLUMN_NAME")
        token_match = next((_PII_TOKENS[token] for token in tokens if token in _PII_TOKENS), None)
        if token_match:
            category, risk = token_match
            if "nome" in tokens and len(tokens) > 1:
                risk = PrivacyRisk.HIGH
            detected.append("COLUMN_NAME")
        for value in values:
            pattern_category = _value_category(str(value))
            if pattern_category is not None:
                category = pattern_category
                risk = PrivacyRisk.HIGH
                detected.append("VALUE_PATTERN")
                break
        if category is None:
            return None
        return PrivacyFinding(
            column=safe_name or name,
            category=category,
            risk=risk,
            detected_by=tuple(sorted(set(detected))),
        )

    def _safe_names(self, names: list[str]) -> tuple[list[str], dict[str, str]]:
        if len(names) > self._limits.max_columns:
            raise ValueError("column count exceeds configured limit")
        if any(not name.strip() or len(name) > self._limits.max_header_length for name in names):
            raise ValueError("dataset header is invalid")
        normalized = [_normalized(name) for name in names]
        if len(set(normalized)) != len(normalized):
            raise ValueError("dataset contains duplicate columns")
        safe: list[str] = []
        mapping: dict[str, str] = {}
        for index, name in enumerate(names, start=1):
            alias = f"column_{index}" if _value_category(name) is not None else name
            safe.append(alias)
            mapping[name] = alias
        return safe, mapping

    def _validate_records(
        self, rows: list[dict[str, Any]]
    ) -> tuple[list[dict[str, Any]], dict[str, str]]:
        if not rows:
            return rows, {}
        names = [str(name) for name in rows[0]]
        safe_names, mapping = self._safe_names(names)
        expected = set(names)
        for row in rows:
            row_names = [str(name) for name in row]
            if set(row_names) != expected or len(row_names) != len(expected):
                raise ValueError("dataset rows have ambiguous width")
        return [
            {safe: row[raw] for raw, safe in zip(names, safe_names, strict=True)} for row in rows
        ], mapping

    def _build(
        self,
        format_name: str,
        rows: list[dict[str, Any]],
        truncated: bool,
        *,
        raw_names: dict[str, str] | None = None,
        encoding: str | None = None,
        delimiter: str | None = None,
        sheets: tuple[str, ...] = (),
        row_estimate: int | None = None,
        extra_findings: tuple[PrivacyFinding, ...] = (),
    ) -> SchemaProfile:
        names = sorted({str(name) for row in rows for name in row})
        reverse = {safe: raw for raw, safe in (raw_names or {}).items()}
        columns: list[ColumnProfile] = []
        candidates: list[str] = []
        findings: list[PrivacyFinding] = []
        for name in names:
            values = [row.get(name) for row in rows]
            non_null = [value for value in values if value not in (None, "")]
            types = sorted({type(value).__name__ for value in non_null})
            distinct = len({repr(value) for value in non_null})
            columns.append(
                ColumnProfile(
                    name=name,
                    type="|".join(types) if types else "unknown",
                    null_rate=(len(values) - len(non_null)) / len(values) if values else 0,
                    distinct_estimate=distinct,
                )
            )
            finding = self._finding(reverse.get(name, name), non_null, name)
            if finding is not None:
                findings.append(finding)
            elif values and len(non_null) == len(values) and distinct == len(values):
                candidates.append(name)
        for finding in extra_findings:
            if finding not in findings:
                findings.append(finding)
        finding_columns = {finding.column for finding in findings}
        candidates = [name for name in candidates if name not in finding_columns]
        warnings = ("row limit reached; estimates are sample-based",) if truncated else ()
        return SchemaProfile(
            format=format_name,
            columns=tuple(columns),
            row_estimate=len(rows) if row_estimate is None else row_estimate,
            encoding=encoding,
            delimiter=delimiter,
            sheets=sheets,
            candidate_keys=tuple(candidates),
            privacy_findings=tuple(findings),
            warnings=warnings,
        )

    def _profile_csv(self, path: Path) -> SchemaProfile:
        payload = path.read_bytes()
        try:
            text = payload.decode("utf-8-sig")
            encoding = "utf-8-sig" if payload.startswith(b"\xef\xbb\xbf") else "utf-8"
        except UnicodeDecodeError:
            text = payload.decode("latin-1")
            encoding = "latin-1"
        try:
            try:
                delimiter = csv.Sniffer().sniff(text[:8192], delimiters=",;\t").delimiter
            except csv.Error:
                non_empty_lines = [line for line in text.splitlines() if line]
                if non_empty_lines and all(
                    not any(candidate in line for candidate in (",", ";", "\t"))
                    for line in non_empty_lines
                ):
                    delimiter = ","
                else:
                    raise
            reader = csv.reader(text.splitlines(), delimiter=delimiter, strict=True)
            header = next(reader, None)
            if header is None:
                raise ValueError("CSV header is invalid")
            safe_names, mapping = self._safe_names(header)
            rows: list[dict[str, Any]] = []
            truncated = False
            for values in reader:
                if len(values) != len(header):
                    raise ValueError("CSV rows have ambiguous width")
                if len(rows) == self._limits.max_rows:
                    truncated = True
                    break
                rows.append(dict(zip(safe_names, values, strict=True)))
        except csv.Error as exc:
            raise ValueError("CSV structure is invalid") from exc
        return self._build(
            "csv",
            rows,
            truncated,
            raw_names=mapping,
            encoding=encoding,
            delimiter=r"\t" if delimiter == "\t" else delimiter,
        )

    @staticmethod
    def _json_object(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
        result: dict[str, Any] = {}
        for key, value in pairs:
            if key in result:
                raise ValueError("JSON object contains duplicate keys")
            result[key] = value
        return result

    def _profile_json(self, path: Path, *, lines: bool) -> SchemaProfile:
        with path.open("r", encoding="utf-8-sig") as stream:
            if lines:
                values = [
                    json.loads(line, object_pairs_hook=self._json_object)
                    for _, line in zip(range(self._limits.max_rows + 1), stream, strict=False)
                    if line.strip()
                ]
            else:
                payload = json.load(stream, object_pairs_hook=self._json_object)
                values = payload if isinstance(payload, list) else [payload]
        if any(not isinstance(value, dict) for value in values):
            raise ValueError("JSON records must be objects")
        truncated = len(values) > self._limits.max_rows
        rows, mapping = self._validate_records(values[: self._limits.max_rows])
        return self._build(
            "jsonl" if lines else "json", rows, truncated, raw_names=mapping, encoding="utf-8"
        )

    def _preflight_xlsx(self, path: Path) -> None:
        try:
            with ZipFile(path) as archive:
                members = archive.infolist()
                if len(members) > self._limits.xlsx_max_members:
                    raise ValueError("XLSX archive exceeds member limit")
                total = 0
                for member in members:
                    total += member.file_size
                    ratio = member.file_size / max(member.compress_size, 1)
                    if (
                        member.file_size > self._limits.xlsx_max_member_uncompressed
                        or ratio > self._limits.xlsx_max_compression_ratio
                    ):
                        raise ValueError("XLSX archive member exceeds safety limits")
                if total > self._limits.xlsx_max_total_uncompressed:
                    raise ValueError("XLSX archive exceeds uncompressed size limit")
        except BadZipFile as exc:
            raise ValueError("XLSX archive is invalid") from exc

    def _sheet_rows(self, worksheet: Any) -> tuple[list[dict[str, Any]], dict[str, str], bool]:
        iterator = worksheet.iter_rows(values_only=True)
        header = next(iterator, None)
        if header is None:
            raise ValueError("XLSX header is invalid")
        names = ["" if value is None else str(value) for value in header]
        safe_names, mapping = self._safe_names(names)
        rows: list[dict[str, Any]] = []
        truncated = False
        for values in iterator:
            if len(values) != len(names):
                raise ValueError("XLSX rows have ambiguous width")
            if len(rows) == self._limits.max_rows:
                truncated = True
                break
            rows.append(dict(zip(safe_names, values, strict=True)))
        return rows, mapping, truncated

    def _profile_xlsx(self, path: Path) -> SchemaProfile:
        self._preflight_xlsx(path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            raw_sheets = tuple(workbook.sheetnames)
            if not raw_sheets:
                raise ValueError("XLSX workbook has no sheets")
            aliases = tuple(f"sheet_{index}" for index in range(1, len(raw_sheets) + 1))
            first_rows: list[dict[str, Any]] = []
            first_mapping: dict[str, str] = {}
            first_truncated = False
            all_findings: list[PrivacyFinding] = []
            for index, raw_sheet in enumerate(raw_sheets):
                rows, mapping, truncated = self._sheet_rows(workbook[raw_sheet])
                profile = self._build("xlsx", rows, truncated, raw_names=mapping)
                all_findings.extend(profile.privacy_findings)
                if index == 0:
                    first_rows, first_mapping, first_truncated = rows, mapping, truncated
            return self._build(
                "xlsx",
                first_rows,
                first_truncated,
                raw_names=first_mapping,
                sheets=aliases,
                extra_findings=tuple(all_findings),
            )
        finally:
            workbook.close()

    def _profile_parquet(self, path: Path) -> SchemaProfile:
        escaped = str(path).replace("'", "''")
        with duckdb.connect(":memory:") as connection:
            schema = connection.execute(
                f"DESCRIBE SELECT * FROM read_parquet('{escaped}')"
            ).fetchall()
            metadata = connection.execute(
                f"SELECT sum(num_rows) FROM parquet_file_metadata('{escaped}')"
            ).fetchone()
            if metadata is None or metadata[0] is None:
                raise RuntimeError("Parquet metadata has no row count")
            total = int(metadata[0])
            sampled = min(total, self._limits.max_rows)
            rows = connection.execute(
                f"SELECT * FROM read_parquet('{escaped}') LIMIT ?", [sampled]
            ).fetchall()
            raw_names = [str(item[0]) for item in schema]
            safe_names, mapping = self._safe_names(raw_names)
            records = [dict(zip(safe_names, row, strict=True)) for row in rows]
        return self._build(
            "parquet", records, sampled < total, raw_names=mapping, row_estimate=total
        )
